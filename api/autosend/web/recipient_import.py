"""Recipient import: turns a CSV upload, Excel upload, or a pasted Google
Sheets / OneDrive / SharePoint link into the list-of-dict row shape the
campaign sender expects.

Split out of campaigns_router.py - this logic doesn't depend on campaign
state and is a natural candidate for reuse anywhere else recipient lists
need to be loaded (e.g. a future non-campaign bulk-import use case).
"""
import base64
import csv
import io
import re

import httpx
import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile

from autosend.web.auth import get_current_web_user

router = APIRouter()

GOOGLE_SHEETS_URL_RE = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)")
ONEDRIVE_DOMAIN_RE = re.compile(r"(1drv\.ms|onedrive\.live\.com|\.sharepoint\.com)", re.IGNORECASE)


def _onedrive_download_url(share_url: str) -> str:
    """Microsoft's documented trick for turning any 'anyone with the link'
    OneDrive/SharePoint share URL into a directly downloadable file, via
    the Shares API's base64 'u!' encoding. Works anonymously for both
    personal OneDrive and SharePoint/business links, no app registration
    or OAuth needed - as long as the link's sharing permission allows it."""
    b64 = base64.urlsafe_b64encode(share_url.encode("utf-8")).decode("utf-8")
    token = "u!" + b64.rstrip("=").replace("/", "_").replace("+", "-")
    return f"https://api.onedrive.com/v1.0/shares/{token}/root/content"


def _clean_phone(value) -> str:
    """Fix Excel's habit of turning phone number columns into floats /
    scientific notation (e.g. 27821234567 -> 2.78212e+10)."""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if "E+" in s.upper():
        s = f"{float(value):.0f}"
    return s

def _parse_csv_bytes(data: bytes):
    encodings_to_try = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    last_error = None
    for enc in encodings_to_try:
        try:
            text = data.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = e
            continue
    raise ValueError(f"Could not decode recipients file with any known encoding: {last_error}")

def _parse_xlsx_bytes(data: bytes, sheet_name: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found in this workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]

    rows = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue
        record = dict(zip(headers, row))
        for key, val in record.items():
            if isinstance(val, (int, float)):
                record[key] = _clean_phone(val)
            elif val is None:
                record[key] = ""
        rows.append(record)
    return rows


async def _fetch_google_sheet_rows(sheets_url: str) -> list[dict]:
    """Only works for sheets shared as 'Anyone with the link can view' -
    exports the sheet as CSV and reuses the CSV parser."""
    match = GOOGLE_SHEETS_URL_RE.search(sheets_url)
    if not match:
        raise HTTPException(status_code=400, detail="Not a valid Google Sheets URL")

    sheet_id = match.group(1)
    gid_match = re.search(r"[#&]gid=(\d+)", sheets_url)
    gid = gid_match.group(1) if gid_match else "0"
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

    async with httpx.AsyncClient(follow_redirects=True) as client:
        resp = await client.get(export_url, timeout=15.0)

    if resp.status_code in (401, 403) or "accounts.google.com" in str(resp.url):
        raise HTTPException(
            status_code=400,
            detail="This Google Sheet isn't publicly viewable. Share it as 'Anyone with the link can view' and try again.",
        )
    resp.raise_for_status()
    return _parse_csv_bytes(resp.content)


def _guess_downloaded_filename(resp: httpx.Response, share_url: str) -> str:
    """Best-effort filename so we know whether to parse the downloaded
    bytes as CSV or Excel."""
    disposition = resp.headers.get("content-disposition", "")
    match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', disposition)
    if match:
        return match.group(1)
    # Fall back to the last path segment of the share URL, if it looks
    # like a filename.
    tail = share_url.split("?")[0].rstrip("/").rsplit("/", 1)[-1]
    if "." in tail:
        return tail
    return "file.xlsx"  # assume Excel if we genuinely can't tell


async def _fetch_onedrive_bytes(share_url: str) -> tuple[bytes, str]:
    download_url = _onedrive_download_url(share_url)
    async with httpx.AsyncClient(follow_redirects=True) as client:
        resp = await client.get(download_url, timeout=20.0)

    if resp.status_code in (401, 403):
        raise HTTPException(
            status_code=400,
            detail="This OneDrive/SharePoint file isn't shared as 'Anyone with the link'. Update the sharing settings and try again.",
        )
    if resp.status_code == 404:
        raise HTTPException(status_code=400, detail="Couldn't find that OneDrive/SharePoint file - check the link is correct and still shared.")
    resp.raise_for_status()

    return resp.content, _guess_downloaded_filename(resp, share_url)


async def _fetch_link_rows(url: str, sheet_name: str | None) -> list[dict]:
    """Dispatches a pasted link to the right fetcher based on domain."""
    if GOOGLE_SHEETS_URL_RE.search(url):
        return await _fetch_google_sheet_rows(url)

    if ONEDRIVE_DOMAIN_RE.search(url):
        data, filename = await _fetch_onedrive_bytes(url)
        if filename.lower().endswith(".csv"):
            return _parse_csv_bytes(data)
        return _parse_xlsx_bytes(data, sheet_name)

    raise HTTPException(status_code=400, detail="That link doesn't look like a Google Sheets or OneDrive/SharePoint link")


async def _load_recipient_rows(recipients_file: UploadFile | None, sheet_link: str | None, sheet_name: str | None = None) -> list[dict]:
    """Single entry point for turning a CSV upload, Excel upload, or a
    pasted Google Sheets / OneDrive / SharePoint link into the same
    list-of-dict row shape the campaign sender already expects."""
    if sheet_link:
        return await _fetch_link_rows(sheet_link, sheet_name)

    if recipients_file is not None and recipients_file.filename:
        data = recipients_file.file.read()
        if recipients_file.filename.lower().endswith((".xlsx", ".xlsm")):
            return _parse_xlsx_bytes(data, sheet_name)
        return _parse_csv_bytes(data)

    raise HTTPException(status_code=400, detail="Provide a recipient CSV/Excel file or a Google Sheets / OneDrive / SharePoint link")


@router.post("/api/campaigns/sheet-names-from-link")
async def list_sheet_names_from_link(sheet_link: str = Form(...), user: dict = Depends(get_current_web_user)):
    """Mirrors /api/campaigns/sheet-names but for a pasted link. Google
    Sheets always exports a single tab as CSV (no picker needed here -
    the gid in the URL already pins the tab). OneDrive/SharePoint links
    that point at an .xlsx get downloaded once to read the sheet index."""
    if GOOGLE_SHEETS_URL_RE.search(sheet_link):
        return {"sheet_names": [], "active_sheet": None}

    if not ONEDRIVE_DOMAIN_RE.search(sheet_link):
        raise HTTPException(status_code=400, detail="That link doesn't look like a Google Sheets or OneDrive/SharePoint link")

    data, filename = await _fetch_onedrive_bytes(sheet_link)
    if filename.lower().endswith(".csv"):
        return {"sheet_names": [], "active_sheet": None}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read that file as Excel - check the link points at a spreadsheet")

    return {"sheet_names": wb.sheetnames, "active_sheet": wb.active.title}


@router.post("/api/campaigns/sheet-names")
def list_sheet_names(recipients_file: UploadFile = File(...), user: dict = Depends(get_current_web_user)):
    """Returns the sheet names in an uploaded .xlsx/.xlsm so the frontend can
    show a picker before the actual campaign submit. Cheap read - doesn't
    load cell values, just the workbook's sheet index."""
    if not recipients_file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Not an Excel file")

    data = recipients_file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read this Excel file - it may be corrupted")

    return {"sheet_names": wb.sheetnames, "active_sheet": wb.active.title}
